import os
import time
import re
import traceback
import threading
import tkinter as tk
from tkinter import scrolledtext
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# --- CONSTANTES GLOBAIS ---
LOG_FILENAME = 'log_validador.log'
EXCEL_FILENAME = 'Extracao_Dados_FSE.xlsx'
INPUT_FILENAME = 'lista.xlsx'
CHROMEDRIVER_PATH = os.path.join(os.getcwd(), "chromedriver.exe")
PORTAL_URL = "https://web.embraer.com.br/irj/portal"

# --- SELETORES SAP (Centralizados para fácil manutenção) ---
# Menus de Navegação
MENU_SUPRIMENTOS = (By.ID, "tabIndex1")
MENU_ORDENS_COMPRA = (By.ID, "L2N0")
MENU_TODAS_ORDENS = (By.ID, "0L3N1")
MENU_DESENHOS_ENGENHARIA = (By.ID, "L2N1")

# Iframes
IFRAME_CONTEUDO_PRINCIPAL = (By.ID, "contentAreaFrame")
IFRAME_CONTEUDO_ANINHADO = (By.XPATH, "//iframe[starts-with(@id, 'ivuFrm_')]")

# Página de Busca de Desenho de Engenharia
CAMPO_BUSCA_PN = (By.XPATH, "//input[contains(@id, 'PartNumber')]")
BOTAO_DESENHO = (By.XPATH, "//a[contains(., 'Desenho')]")
BOTAO_VOLTAR = (By.XPATH, "//a[contains(., 'Voltar')]")
CAMPO_RESULTADO_REV = (By.XPATH, "//span[contains(text(), 'Rev ')]")

# Página de Busca FSE
CAMPO_BUSCA_OC = (By.XPATH, "//input[@ng-model='vm.search.orderNumber']")
CAMPO_BUSCA_ITEM = (By.XPATH, "//input[@ng-model='vm.search.orderLine']")
BOTAO_BUSCAR_FSE = (By.ID, "searchBtn")
BOTAO_DETALHES_FSE = (By.XPATH, "//button[contains(@ng-click, 'vm.showFseDetails')]")
HEADER_FSE = (By.ID, "fseHeader")

# Seletores de Dados FSE
DADO_OC_ITEM = (By.XPATH, "//*[@id='fseHeader']/div[1]/div[5]")
DADO_CODEM_DATA = (By.XPATH, "//*[@id='fseHeader']/div[3]/div[1]")
DADO_PN_REV_LID = (By.XPATH, "//*[@id='fseHeader']/div[3]/div[2]")
DADO_IND_RASTR = (By.XPATH, "//*[@id='fseHeader']/div[2]/div[3]")
DADO_SERIAIS = (By.XPATH, "//*[text()='NÚMERO DE SERIAÇÃO']/following-sibling::div//span")


class ValidadorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Validador de Revisão de Engenharia v2.0")
        self.root.geometry("850x650")
        self.root.attributes('-topmost', True)
        
        self.user_action_event = threading.Event()
        self.driver = None

        main_frame = tk.Frame(root, padx=10, pady=10)
        main_frame.pack(expand=True, fill='both')

        top_frame = tk.Frame(main_frame)
        top_frame.pack(fill='x', pady=(0, 5))

        self.label_status = tk.Label(top_frame, text="Pronto para iniciar.", font=("Helvetica", 12, "bold"), fg="#00529B", pady=10, wraplength=700, justify='center')
        self.label_status.pack()

        self.action_button = tk.Button(top_frame, text="Iniciar Automação", command=self.iniciar_automacao_thread, font=("Helvetica", 12, "bold"), bg="#4CAF50", fg="white", padx=20, pady=10)
        self.action_button.pack(pady=(5, 10))

        log_label = tk.Label(main_frame, text="Log em Tempo Real:", font=("Helvetica", 10, "bold"))
        log_label.pack(fill='x', pady=(10, 0))
        self.log_text = scrolledtext.ScrolledText(main_frame, state='disabled', wrap=tk.WORD, font=("Courier New", 9))
        self.log_text.pack(expand=True, fill='both', pady=5)
        
        self.log_path = os.path.join(os.getcwd(), LOG_FILENAME)
        self.excel_path = os.path.join(os.getcwd(), EXCEL_FILENAME)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def on_closing(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
        self.root.destroy()

    def registrar_log(self, mensagem):
        log_entry = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {mensagem}\n"
        with open(self.log_path, 'a', encoding='utf-8') as log_file:
            log_file.write(log_entry)
        
        def update_gui():
            self.log_text.config(state='normal')
            self.log_text.insert(tk.END, log_entry)
            self.log_text.see(tk.END)
            self.log_text.config(state='disabled')
        self.root.after(0, update_gui)

    def update_status(self, text, color="#00529B"):
        self.root.after(0, lambda: self.label_status.config(text=text, fg=color))

    def iniciar_automacao_thread(self):
        self.action_button.config(state='disabled', text="Executando...")
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        
        threading.Thread(target=self.run_automation, daemon=True).start()

    def prompt_user_action(self, message):
        self.user_action_event.clear()
        self.root.after(0, lambda: [
            self.update_status(message, color="#E69500"),
            self.action_button.config(text="Continuar", command=self.signal_user_action, state="normal")
        ])
        self.user_action_event.wait()
        self.root.after(0, lambda: self.action_button.config(state='disabled', text="Executando..."))

    def signal_user_action(self):
        self.user_action_event.set()
        
    def setup_driver(self):
        """Inicializa e retorna uma instância do WebDriver."""
        if not self.driver:
            self.update_status("Configurando o navegador...")
            service = ChromeService(executable_path=CHROMEDRIVER_PATH)
            options = webdriver.ChromeOptions()
            options.add_argument("--start-maximized")
            # options.add_argument("--headless") # Descomente para rodar em segundo plano
            self.driver = webdriver.Chrome(service=service, options=options)
            self.registrar_log("Navegador iniciado.")
        return self.driver, WebDriverWait(self.driver, 20)

    def setup_excel(self):
        if not os.path.exists(self.excel_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Dados FSE"
            self.headers = [
                "OS", "OC / Item", "CODEM / DT. REV. ROT.", "PN / REV. PN / LID", 
                "IND. RASTR.", "NÚMERO DE SERIAÇÃO", "PN extraído", "REV. FSE",
                "REV. Engenharia", "Status Comparação"
            ]
            sheet.append(self.headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            workbook.save(self.excel_path)
            self.registrar_log(f"Arquivo Excel '{EXCEL_FILENAME}' criado.")
        else:
            workbook = openpyxl.load_workbook(self.excel_path)
            sheet = workbook.active
            self.headers = [cell.value for cell in sheet[1]]
            self.registrar_log(f"Arquivo Excel '{EXCEL_FILENAME}' já existe. Será atualizado.")

    def run_automation(self):
        try:
            self.update_status("Iniciando automação...")
            self.setup_excel()

            df_input = pd.read_excel(INPUT_FILENAME, sheet_name='baixar_lm', engine='openpyxl')
            df_input.rename(columns={df_input.columns[0]: 'OS', df_input.columns[1]: 'OC_COMPLETA'}, inplace=True)
            df_input[['OC_antes', 'OC_depois']] = df_input['OC_COMPLETA'].astype(str).str.split('/', expand=True, n=1)
            df_input['OS'] = df_input['OS'].astype(str)
            self.registrar_log(f"Arquivo '{INPUT_FILENAME}' lido com {len(df_input)} itens.")

            os_ja_extraidas = set()
            try:
                df_existente = pd.read_excel(self.excel_path)
                if 'OS' in df_existente.columns:
                    os_ja_extraidas = set(df_existente['OS'].astype(str))
                self.registrar_log(f"{len(os_ja_extraidas)} OSs já encontradas no arquivo de resultados.")
            except FileNotFoundError:
                self.registrar_log("Arquivo de resultados não encontrado. Iniciando do zero.")
            
            df_a_extrair = df_input[~df_input['OS'].isin(os_ja_extraidas)].copy()

            if not df_a_extrair.empty:
                self.executar_etapa_extracao(df_a_extrair)
            else:
                self.registrar_log("Nenhuma nova OS para extrair. Pulando para a Etapa de Comparação.")

            self.executar_etapa_comparacao(set(df_input['OS']))

            self.update_status("Processo concluído com sucesso!", "#008A00")
        except Exception as e:
            error_details = traceback.format_exc()
            self.registrar_log(f"ERRO CRÍTICO: {error_details}")
            self.update_status(f"Erro Crítico: {e}", "red")
        finally:
            if self.driver:
                self.registrar_log("Fechando navegador.")
                self.driver.quit()
                self.driver = None
            self.root.after(0, lambda: self.action_button.pack_forget())

    def executar_etapa_extracao(self, df_a_extrair):
        self.update_status(f"ETAPA 1: Extraindo dados de {len(df_a_extrair)} novas OSs...")
        driver, wait = self.setup_driver()
        driver.get(PORTAL_URL)
        self.prompt_user_action("Faça o login no portal. Quando a página principal carregar, clique em 'Continuar'.")
        
        self.navegar_para_fse_busca(wait)

        for index, row in df_a_extrair.iterrows():
            os_num = str(row['OS'])
            self.update_status(f"Extraindo dados da OS: {os_num} ({index + 1}/{len(df_a_extrair)})...")
            dados_fse = self.extrair_dados_fse(wait, os_num, row['OC_antes'], row['OC_depois'])
            if dados_fse:
                workbook = openpyxl.load_workbook(self.excel_path)
                sheet = workbook.active
                sheet.append(list(dados_fse.values()))
                workbook.save(self.excel_path)
        self.registrar_log("Etapa 1 (Extração) concluída.")
    
    def executar_etapa_comparacao(self, os_relevantes):
        self.update_status("ETAPA 2: Verificando comparações pendentes...")
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook.active
        col_indices = {name: i + 1 for i, name in enumerate(self.headers)}

        linhas_a_comparar = []
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            os_da_linha = str(row[col_indices["OS"] - 1])
            status_comp = row[col_indices["Status Comparação"] - 1]
            if os_da_linha in os_relevantes and not status_comp:
                linhas_a_comparar.append(i + 2)

        if not linhas_a_comparar:
            self.registrar_log("Nenhuma comparação pendente para os itens da lista.")
            return

        self.update_status(f"ETAPA 2: Comparando {len(linhas_a_comparar)} itens...")
        driver, wait = self.setup_driver()
        if "irj/portal" not in driver.current_url:
             driver.get(PORTAL_URL)
             self.prompt_user_action("Faça o login novamente para a etapa de comparação e clique em 'Continuar'.")

        self.navegar_para_desenhos_engenharia(wait)

        for row_num in linhas_a_comparar:
            row_cells = sheet[row_num]
            pn = row_cells[col_indices["PN extraído"] - 1].value
            rev_fse = row_cells[col_indices["REV. FSE"] - 1].value
            self.update_status(f"Comparando PN: {pn} (linha {row_num})...")

            if pn and pn != "Não encontrado":
                rev_eng = self.buscar_revisao_engenharia(wait, pn)
                status = "DIVERGENTE" if rev_eng != "Não encontrada" and rev_fse != rev_eng else "OK"
                if rev_eng == "Não encontrada": status = "FALHA NA BUSCA"
                
                sheet.cell(row=row_num, column=col_indices["REV. Engenharia"], value=rev_eng)
                sheet.cell(row=row_num, column=col_indices["Status Comparação"], value=status)
            else:
                sheet.cell(row=row_num, column=col_indices["Status Comparação"], value="PN NÃO ENCONTRADO NA FSE")
            
            workbook.save(self.excel_path)
        self.registrar_log("Etapa 2 (Comparação) concluída.")

    def navegar_para_fse_busca(self, wait):
        original_window = self.driver.current_window_handle
        wait.until(EC.element_to_be_clickable(MENU_ORDENS_COMPRA)).click()
        wait.until(EC.element_to_be_clickable(MENU_TODAS_ORDENS)).click() # Exemplo
        # ...adicionar cliques para chegar na busca FSE
        
        # A lógica para alternar de janela/aba pode ser necessária aqui
        # wait.until(EC.number_of_windows_to_be(2))
        # for handle in self.driver.window_handles:
        #     if handle != original_window:
        #         self.driver.switch_to.window(handle)
        #         break
        
        self.prompt_user_action("Navegue manualmente até 'Busca FSe'. Quando a tela carregar, clique em 'Continuar'.")

    def extrair_dados_fse(self, wait, os_num, oc, item):
        try:
            self.registrar_log(f"Buscando OS: {os_num} | OC: {oc}/{item}")
            wait.until(EC.visibility_of_element_located(CAMPO_BUSCA_OC)).clear()
            self.driver.find_element(*CAMPO_BUSCA_OC).send_keys(oc)
            wait.until(EC.visibility_of_element_located(CAMPO_BUSCA_ITEM)).clear()
            self.driver.find_element(*CAMPO_BUSCA_ITEM).send_keys(item)
            wait.until(EC.element_to_be_clickable(BOTAO_BUSCAR_FSE)).click()
            wait.until(EC.element_to_be_clickable(BOTAO_DETALHES_FSE)).click()
            
            wait.until(EC.visibility_of_element_located(HEADER_FSE))
            
            dados = {"OS": os_num}
            dados["OC / Item"] = self.safe_find_text(*DADO_OC_ITEM).replace('\n', ' ')
            dados["CODEM / DT. REV. ROT."] = self.safe_find_text(*DADO_CODEM_DATA).replace('CODEM / DT. REV. ROT.\n', '').replace('\n', ' | ')
            dados["PN / REV. PN / LID"] = self.safe_find_text(*DADO_PN_REV_LID).replace('PN / REV. PN / LID\n', '').replace('\n', ' | ')
            dados["IND. RASTR."] = self.safe_find_text(*DADO_IND_RASTR).replace('IND. RASTR.\n', '').strip()
            
            seriais = self.driver.find_elements(*DADO_SERIAIS)
            dados["NÚMERO DE SERIAÇÃO"] = ", ".join([el.text for el in seriais if el.text.strip()])

            pn_rev_raw = dados["PN / REV. PN / LID"]
            pn_match = re.search(r'(\d+-\d+-\d+)', pn_rev_raw)
            rev_match = re.search(r'\b([A-Z])\b', pn_rev_raw) # Busca uma letra maiúscula isolada
            
            dados["PN extraído"] = pn_match.group(1) if pn_match else "Não encontrado"
            dados["REV. FSE"] = rev_match.group(1) if rev_match else "Não encontrada"
            dados["REV. Engenharia"] = ""
            dados["Status Comparação"] = ""

            self.driver.back() # Volta para a tela de busca
            wait.until(EC.visibility_of_element_located(CAMPO_BUSCA_OC))
            return dados

        except Exception as e:
            self.registrar_log(f"ERRO ao extrair dados da OS {os_num}: {e}")
            self.tirar_print_de_erro(os_num, "extracao_FSE")
            self.driver.get(self.driver.current_url) # Recarrega a página de busca
            return None

    def navegar_para_desenhos_engenharia(self, wait):
        self.driver.switch_to.window(self.driver.window_handles[0])
        if "irj/portal" not in self.driver.current_url:
            self.driver.get(PORTAL_URL)
        wait.until(EC.element_to_be_clickable(MENU_DESENHOS_ENGENHARIA)).click()
        self.prompt_user_action("Valide se a tela 'Desenhos Engenharia' está aberta e clique em 'Continuar'.")

    def buscar_revisao_engenharia(self, wait, part_number):
        self.registrar_log(f"Buscando revisão para o PN: {part_number}")
        try:
            wait.until(EC.frame_to_be_available_and_switch_to_it(IFRAME_CONTEUDO_PRINCIPAL))
            wait.until(EC.frame_to_be_available_and_switch_to_it(IFRAME_CONTEUDO_ANINHADO))
            
            campo_pn = wait.until(EC.visibility_of_element_located(CAMPO_BUSCA_PN))
            campo_pn.clear()
            campo_pn.send_keys(part_number)

            botao_desenho = wait.until(EC.element_to_be_clickable(BOTAO_DESENHO))
            self.driver.execute_script("arguments[0].click();", botao_desenho)

            rev_element = wait.until(EC.visibility_of_element_located(CAMPO_RESULTADO_REV))
            revisao = rev_element.text.split(" ")[-1]
            self.registrar_log(f"SUCESSO: Revisão encontrada para PN {part_number}: {revisao}")
            
            botao_voltar = wait.until(EC.element_to_be_clickable(BOTAO_VOLTAR))
            self.driver.execute_script("arguments[0].click();", botao_voltar)
            wait.until(EC.visibility_of_element_located(CAMPO_BUSCA_PN)) # Confirma retorno

            return revisao
        except Exception:
            self.registrar_log(f"ERRO: Revisão não encontrada para o PN {part_number}.")
            self.tirar_print_de_erro(part_number, "busca_revisao")
            return "Não encontrada"
        finally:
            self.driver.switch_to.default_content()

    def safe_find_text(self, by, value):
        try:
            return self.driver.find_element(by, value).text
        except NoSuchElementException:
            return ""

    def tirar_print_de_erro(self, identificador, etapa):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_screenshot = f"erro_{etapa}_{identificador}_{timestamp}.png"
        try:
            self.driver.save_screenshot(nome_screenshot)
            self.registrar_log(f"Screenshot de erro salvo: '{nome_screenshot}'")
        except Exception as e:
            self.registrar_log(f"FALHA AO SALVAR SCREENSHOT: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ValidadorGUI(root)
    root.mainloop()