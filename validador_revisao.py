import os
import time
import re
import traceback
import tkinter as tk
from tkinter import scrolledtext
import threading
from datetime import datetime
import pandas as pd
import openpyxl
import pyodbc

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains

# --- CONSTANTES GLOBAIS ---
LOG_FILENAME = 'log_validador.txt'
EXCEL_FILENAME = 'Extracao_Dados_FSE.xlsx'
ERROS_DIR = 'erros'

class ValidadorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Validador de Revisão de Engenharia")
        self.root.geometry("850x650")
        self.root.attributes('-topmost', True)
        
        self.user_action_event = threading.Event()
        self.stop_event = threading.Event()
        self.reprocess_choice = "" 
        self.driver = None

        main_frame = tk.Frame(root, padx=10, pady=10)
        main_frame.pack(expand=True, fill='both')

        top_frame = tk.Frame(main_frame)
        top_frame.pack(fill='x', pady=(0, 5))

        self.label_status = tk.Label(top_frame, text="Pronto para iniciar.", font=("Helvetica", 12, "bold"), fg="#00529B", pady=10, wraplength=700, justify='center')
        self.label_status.pack()

        self.action_frame = tk.Frame(top_frame)
        self.action_frame.pack(pady=(5,10))

        self.action_button = tk.Button(self.action_frame, text="Iniciar Automação", command=self.iniciar_automacao_thread, font=("Helvetica", 12, "bold"), bg="#4CAF50", fg="white", padx=20, pady=10)
        self.action_button.pack(side='left', padx=5)

        self.stop_button = tk.Button(self.action_frame, text="Parar Automação", command=self.request_stop, font=("Helvetica", 12, "bold"), bg="#f44336", fg="white", padx=20, pady=10)

        self.reprocess_frame = tk.Frame(main_frame)
        self.reprocess_button = tk.Button(self.reprocess_frame, text="Reprocessar Itens com Erro", command=lambda: self.set_reprocess_choice("reprocess"), font=("Helvetica", 10, "bold"), bg="#FFA500", fg="white")
        self.finish_button = tk.Button(self.reprocess_frame, text="Finalizar", command=lambda: self.set_reprocess_choice("finish"), font=("Helvetica", 10))
        self.reprocess_button.pack(side='left', padx=5)
        self.finish_button.pack(side='left', padx=5)
        self.reprocess_frame.pack_forget()

        log_label = tk.Label(main_frame, text="Log de Atividades:", font=("Helvetica", 10, "bold"))
        log_label.pack(fill='x', pady=(10, 0))
        self.log_text = scrolledtext.ScrolledText(main_frame, state='disabled', wrap=tk.WORD, font=("Courier New", 9))
        self.log_text.pack(expand=True, fill='both', pady=5)
        
        self.log_path = os.path.join(os.getcwd(), LOG_FILENAME)
        self.excel_path = os.path.join(os.getcwd(), EXCEL_FILENAME)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def on_closing(self):
        self.stop_event.set()
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
        self.root.destroy()

    def registrar_log(self, mensagem):
        log_entry = f"[{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}] {mensagem}\n"
        with open(self.log_path, 'a', encoding='utf-8') as log_file:
            log_file.write(log_entry)
        
        def update_gui():
            self.log_text.config(state='normal')
            self.log_text.insert(tk.END, log_entry)
            self.log_text.see(tk.END)
            self.log_text.config(state='disabled')
        self.root.after(0, update_gui)

    def update_status(self, text, color="#00529B"):
        self.root.after(0, lambda: self.label_status.config(text=text, fg=color))

    def iniciar_automacao_thread(self):
        self.stop_event.clear()
        self.action_button.pack_forget()
        self.stop_button.pack(side='left', padx=5)
        self.stop_button.config(state='normal')
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        
        threading.Thread(target=self.run_automation_loop, daemon=True).start()

    def request_stop(self):
        self.registrar_log("Solicitação de parada recebida... A automação será interrompida após a tarefa atual.")
        self.update_status("Parando automação...", "#E69500")
        self.stop_event.set()
        self.stop_button.config(state='disabled')

    def prompt_user_action(self, message):
        self.user_action_event.clear()
        self.root.after(0, lambda: [
            self.update_status(message, color="#E69500"),
            self.stop_button.pack_forget(),
            self.action_button.config(text="Continuar", command=self.signal_user_action, state="normal"),
            self.action_button.pack(side='left', padx=5)
        ])
        self.user_action_event.wait()
        self.root.after(0, lambda: [
            self.action_button.pack_forget(),
            self.stop_button.pack(side='left', padx=5)
        ])

    def signal_user_action(self):
        self.user_action_event.set()
        
    def set_reprocess_choice(self, choice):
        self.reprocess_choice = choice

    def setup_excel(self):
        if not os.path.exists(self.excel_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Dados FSE"
            self.headers = [
                "OS", "OC", "Item", "CODEM", "DT. REV. ROT.", "PN", "REV. PN", "LID", "PLANTA",
                "IND. RASTR.", "NÚMERO DE SERIAÇÃO", "PN extraído", 
                "REV. FSE", "REV. Engenharia", "Revisão do Banco",
                "Status (Eng vs FSE)", "Detalhes (Eng vs FSE)",
                "Status (Banco vs FSE)", "Detalhes (Banco vs FSE)",
                "Status (Banco vs Eng)", "Detalhes (Banco vs Eng)",
                "COD_OS_COMPLETO", "COD_PECAS", "N_OS_CLIENTE", "N_DESENHO", "QTDE_PECAS", "DT_ENTRADA",
                "U_ZLT_REVISAO_2D", "U_ZLT_REVISAO_DES_3D", "U_ZLT_REVISAO_DI", "U_ZLT_REVISAO_DI_F2",
                "U_ZLT_REVISAO_DI_F3", "U_ZLT_REVISAO_FI", "U_ZLT_REVISAO_LP", "U_ZLT_REVISAO_LP_F2",
                "U_ZLT_REVISAO_LP_F3", "U_ZLT_REVISAO_PN", "U_ZLT_REVISAO_ROT", "U_REVISAO_3D_MF",
                "U_CLASSE_ELEB", "U_PN_DE_PROJETO", "U_OM", "CAMPO_ADICIONAL9"
            ]
            sheet.append(self.headers)
            for cell in sheet[1]:
                cell.font = openpyxl.styles.Font(bold=True)
            workbook.save(self.excel_path)
            self.registrar_log(f"Arquivo de resultados '{EXCEL_FILENAME}' criado com sucesso.")
        else:
            workbook = openpyxl.load_workbook(self.excel_path)
            sheet = workbook.active
            self.headers = [cell.value for cell in sheet[1]]
            self.registrar_log(f"Arquivo de resultados '{EXCEL_FILENAME}' encontrado.")
    
    def consultar_dados_banco(self, part_number):
        self.registrar_log(f"Consultando banco de dados para o PN: {part_number}...")
        
        string_conexao = (
            r'DRIVER={SQL Server};'
            r'SERVER=172.20.1.7;DATABASE=CPS;UID=sa;PWD=masterkey;'
        )
        
        comando_sql = """
            SELECT  
              COD_OS_COMPLETO, COD_PECAS, N_OS_CLIENTE, N_DESENHO, QTDE_PECAS, DT_ENTRADA,
              U_ZLT_REVISAO_2D, U_ZLT_REVISAO_DES_3D, U_ZLT_REVISAO_DI, U_ZLT_REVISAO_DI_F2,
              U_ZLT_REVISAO_DI_F3, U_ZLT_REVISAO_FI, U_ZLT_REVISAO_LP, U_ZLT_REVISAO_LP_F2,
              U_ZLT_REVISAO_LP_F3, U_ZLT_REVISAO_PN, U_ZLT_REVISAO_ROT, U_REVISAO_3D_MF,
              U_CLASSE_ELEB, U_PN_DE_PROJETO, U_OM, CAMPO_ADICIONAL9
            FROM TOS_AUX
            WHERE N_DESENHO = ? 
        """
        
        colunas_db = [
            "COD_OS_COMPLETO", "COD_PECAS", "N_OS_CLIENTE", "N_DESENHO", "QTDE_PECAS", "DT_ENTRADA",
            "U_ZLT_REVISAO_2D", "U_ZLT_REVISAO_DES_3D", "U_ZLT_REVISAO_DI", "U_ZLT_REVISAO_DI_F2",
            "U_ZLT_REVISAO_DI_F3", "U_ZLT_REVISAO_FI", "U_ZLT_REVISAO_LP", "U_ZLT_REVISAO_LP_F2",
            "U_ZLT_REVISAO_LP_F3", "U_ZLT_REVISAO_PN", "U_ZLT_REVISAO_ROT", "U_REVISAO_3D_MF",
            "U_CLASSE_ELEB", "U_PN_DE_PROJETO", "U_OM", "CAMPO_ADICIONAL9"
        ]
        
        try:
            with pyodbc.connect(string_conexao, timeout=5) as conexao:
                cursor = conexao.cursor()
                cursor.execute(comando_sql, part_number)
                resultado = cursor.fetchone()
                
                if resultado:
                    self.registrar_log(f"Dados encontrados no banco para o PN: {part_number}")
                    return dict(zip(colunas_db, resultado))
                else:
                    self.registrar_log(f"PN {part_number} não encontrado no banco de dados.")
                    return {col: "Não encontrado no BD" for col in colunas_db}
        except Exception as e:
            self.registrar_log(f"ERRO ao consultar o banco de dados: {e}")
            return {col: "Erro no BD" for col in colunas_db}

    def run_automation_loop(self):
        reprocess_mode = False
        while not self.stop_event.is_set():
            try:
                # Carrega a lista de OSs a serem processadas (a original ou a de erros)
                if not reprocess_mode:
                    df_input = pd.read_excel('lista.xlsx', sheet_name='lista', engine='openpyxl')
                    df_input.rename(columns={df_input.columns[0]: 'OS'}, inplace=True)
                    df_input[['OC_antes', 'OC_depois']] = df_input.iloc[:, 1].astype(str).str.split('/', expand=True, n=1)
                    df_input['OS'] = df_input['OS'].astype(str)
                
                # Executa o ciclo principal de extração e comparação
                self.run_automation_cycle(df_input)
                
                if self.stop_event.is_set(): break
                
                # Verifica se há erros para reprocessar
                self.update_status("Verificando se existem erros para reprocessar...", "#00529B")
                erros_df = self.check_for_errors()
                
                if not erros_df.empty:
                    self.reprocess_choice = ""
                    self.root.after(0, lambda: [
                        self.update_status(f"{len(erros_df)} OSs com falhas encontradas. Deseja reprocessá-las?", "#E69500"),
                        self.action_frame.pack_forget(),
                        self.reprocess_frame.pack()
                    ])
                    
                    while not self.reprocess_choice and not self.stop_event.is_set():
                        time.sleep(0.1)

                    self.root.after(0, self.reprocess_frame.pack_forget)
                    
                    if self.reprocess_choice == "reprocess":
                        self.registrar_log(f"--- INICIANDO REPROCESSO PARA {len(erros_df)} OSs COM FALHA ---")
                        df_input = erros_df
                        reprocess_mode = True
                        self.clear_error_status_in_excel(erros_df) # Limpa status no Excel
                        continue # Volta para o início do loop com a lista de erros
                    else:
                        self.update_status("Processo finalizado com itens pendentes.", "#00529B")
                        break # Encerra o loop se o usuário escolher "Finalizar"
                else:
                    self.update_status("Processo concluído com sucesso e sem erros!", "#008A00")
                    break # Encerra o loop se não houver erros

            except Exception as e:
                error_details = traceback.format_exc()
                self.registrar_log(f"ERRO CRÍTICO NO LOOP PRINCIPAL: {error_details}")
                self.update_status(f"Erro Crítico: {e}", "red")
                break
            finally:
                if self.driver:
                    self.driver.quit()
                    self.driver = None
        
        # Garante que a interface volte ao estado inicial
        self.root.after(0, lambda: [
            self.stop_button.pack_forget(),
            self.reprocess_frame.pack_forget(),
            self.action_button.config(state='normal', text="Iniciar Automação"),
            self.action_button.pack(side='left', padx=5),
            self.action_frame.pack()
        ])
        self.registrar_log("--- FIM DA EXECUÇÃO ---")

    def check_for_errors(self):
        """Lê o Excel e retorna um DataFrame com as linhas que contêm erros."""
        df_results = pd.read_excel(self.excel_path)
        df_results['OS'] = df_results['OS'].astype(str)
        
        # Considera erro se qualquer um dos status não for "OK"
        falha = ~df_results['Status (Eng vs FSE)'].astype(str).str.contains("OK", na=False)
        
        df_erros = df_results[falha].copy()
        
        if not df_erros.empty:
            df_erros[['OC_antes', 'OC_depois']] = df_erros['OC'].astype(str).str.split('/', expand=True, n=1)
        
        return df_erros

    def clear_error_status_in_excel(self, df_erros):
        """Apaga o status de erro no Excel para as OSs que serão reprocessadas."""
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook.active
        col_indices = {name: i + 1 for i, name in enumerate(self.headers)}
        os_erros = set(df_erros['OS'].astype(str))

        for row_idx in range(2, sheet.max_row + 1):
            os_atual = str(sheet.cell(row=row_idx, column=col_indices["OS"]).value)
            if os_atual in os_erros:
                # Limpa todas as colunas de status e detalhes para forçar o reprocessamento completo
                for col_name in ["REV. Engenharia", "Revisão do Banco", "Status (Eng vs FSE)", "Detalhes (Eng vs FSE)", "Status (Banco vs FSE)", "Detalhes (Banco vs FSE)", "Status (Banco vs Eng)", "Detalhes (Banco vs Eng)"]:
                    sheet.cell(row=row_idx, column=col_indices[col_name], value=None)
        
        workbook.save(self.excel_path)
        self.registrar_log(f"Limpando status de {len(os_erros)} OSs para reprocessamento.")


    def run_automation_cycle(self, df_to_process):
        self.registrar_log("--- INÍCIO DO CICLO DE EXECUÇÃO ---")
        self.setup_excel()
        
        # ETAPA 1: EXTRAÇÃO DE DADOS DA FSE
        os_ja_extraidas = set()
        try:
            df_existente = pd.read_excel(self.excel_path)
            if 'OS' in df_existente.columns:
                os_ja_extraidas = set(df_existente['OS'].astype(str))
        except FileNotFoundError:
            pass
        
        df_a_extrair = df_to_process[~df_to_process['OS'].isin(os_ja_extraidas)].copy()
        
        if not df_a_extrair.empty:
            if self.stop_event.is_set(): return
            
            if not self.driver:
                self.update_status("Configurando navegador...")
                self._setup_driver()
                self.driver.get("https://web.embraer.com.br/irj/portal")
                self.prompt_user_action("Por favor, faça o login no portal.")
            
            wait = WebDriverWait(self.driver, 20)
            self.update_status(f"Extraindo dados de {len(df_a_extrair)} Fichas Seguidoras...")
            self.navegar_para_fse_busca(wait)
            
            novas_linhas_dados = []
            for _, row in df_a_extrair.iterrows():
                if self.stop_event.is_set(): break
                dados_fse = self.extrair_dados_fse(wait, str(row['OS']), row['OC_antes'], row['OC_depois'])
                if dados_fse: novas_linhas_dados.append(dados_fse)
            
            if novas_linhas_dados:
                workbook = openpyxl.load_workbook(self.excel_path)
                sheet = workbook.active
                for dados in novas_linhas_dados:
                    sheet.append([dados.get(h, "") for h in self.headers])
                workbook.save(self.excel_path)
                self.registrar_log(f"Salvos {len(novas_linhas_dados)} novos registros de FSE.")

        if self.stop_event.is_set(): return

        # ETAPA 2: COMPARAÇÃO DE REVISÕES
        self.update_status("Etapa 2: Verificando e comparando revisões...")
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook.active
        col_indices = {name: i+1 for i, name in enumerate(self.headers)}
        
        df_para_comparar_full = pd.read_excel(self.excel_path)
        df_para_comparar_full['OS'] = df_para_comparar_full['OS'].astype(str)
        # Filtra apenas as OSs do ciclo atual que precisam de comparação
        df_ciclo_atual = df_para_comparar_full[df_para_comparar_full['OS'].isin(df_to_process['OS'].tolist())]
        linhas_a_comparar = df_ciclo_atual[df_ciclo_atual['Status (Eng vs FSE)'].isna()].index + 2
        
        if not linhas_a_comparar.empty:
            self.registrar_log(f"Encontradas {len(linhas_a_comparar)} OCs para comparar.")
            if not self.driver:
                self.update_status("Configurando navegador para Etapa 2...")
                self._setup_driver()
                self.driver.get("https://web.embraer.com.br/irj/portal")
                self.prompt_user_action("Faça o login para a comparação e clique em 'Continuar'.")

            wait = WebDriverWait(self.driver, 20)
            self.navegar_para_desenhos_engenharia(wait)

            for row_num in linhas_a_comparar:
                if self.stop_event.is_set(): break
                pn_extraido = sheet.cell(row=row_num, column=col_indices["PN extraído"]).value
                rev_fse = sheet.cell(row=row_num, column=col_indices["REV. FSE"]).value
                
                if pn_extraido and pn_extraido != "Não encontrado":
                    self.update_status(f"Comparando PN: {pn_extraido}...")
                    rev_engenharia = self.buscar_revisao_engenharia(wait, pn_extraido)
                    dados_banco = self.consultar_dados_banco(pn_extraido)
                    revisao_banco = dados_banco.get("U_ZLT_REVISAO_PN", "Chave não encontrada")
                    
                    status_eng_fse, detalhes_eng_fse = self.comparar_revisoes(rev_engenharia, rev_fse, "ENG", "FSE")
                    status_banco_fse, detalhes_banco_fse = self.comparar_revisoes(revisao_banco, rev_fse, "BANCO", "FSE")
                    status_banco_eng, detalhes_banco_eng = self.comparar_revisoes(revisao_banco, rev_engenharia, "BANCO", "ENG")

                    # Atualiza a planilha em memória
                    sheet.cell(row=row_num, column=col_indices["REV. Engenharia"], value=rev_engenharia)
                    sheet.cell(row=row_num, column=col_indices["Revisão do Banco"], value=revisao_banco)
                    sheet.cell(row=row_num, column=col_indices["Status (Eng vs FSE)"], value=status_eng_fse)
                    sheet.cell(row=row_num, column=col_indices["Detalhes (Eng vs FSE)"], value=detalhes_eng_fse)
                    sheet.cell(row=row_num, column=col_indices["Status (Banco vs FSE)"], value=status_banco_fse)
                    sheet.cell(row=row_num, column=col_indices["Detalhes (Banco vs FSE)"], value=detalhes_banco_fse)
                    sheet.cell(row=row_num, column=col_indices["Status (Banco vs Eng)"], value=status_banco_eng)
                    sheet.cell(row=row_num, column=col_indices["Detalhes (Banco vs Eng)"], value=detalhes_banco_eng)
                    
                    for nome_coluna, valor in dados_banco.items():
                        if nome_coluna in col_indices:
                            sheet.cell(row=row_num, column=col_indices[nome_coluna], value=valor)
                else:
                    sheet.cell(row=row_num, column=col_indices["Status (Eng vs FSE)"], value="PN NÃO ENCONTRADO NA FSE")

            self.registrar_log("Salvando resultados das comparações no Excel...")
            workbook.save(self.excel_path)
            self.registrar_log("Arquivo salvo com sucesso.")
        else:
             self.registrar_log("Nenhuma comparação pendente para este ciclo.")

    def _setup_driver(self):
        """Inicializa o objeto do WebDriver."""
        caminho_chromedriver = os.path.join(os.getcwd(), "chromedriver.exe")
        service = ChromeService(executable_path=caminho_chromedriver)
        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument("--disable-blink-features=AutomationControlled")
        self.driver = webdriver.Chrome(service=service, options=options)

    def comparar_revisoes(self, rev1, rev2, nome1, nome2):
        detalhes = f"{nome1}: {rev1} vs {nome2}: {rev2}"
        # Assume falha até que os dados sejam validados
        status = "FALHA"
        is_rev1_valida = rev1 and "Não encontrad" not in str(rev1) and "Erro" not in str(rev1) and "Falha" not in str(rev1)
        is_rev2_valida = rev2 and "Não encontrad" not in str(rev2) and "Erro" not in str(rev2) and "Falha" not in str(rev2)
        
        if is_rev1_valida and is_rev2_valida:
            if str(rev1).strip().upper() == str(rev2).strip().upper():
                status = "OK"
            else:
                status = "DIVERGENTE"
        
        return status, detalhes

    def extrair_dados_fse(self, wait, os_num, oc1, oc2):
        oc_completa = f"{oc1}/{oc2}"
        try:
            self.update_status(f"Extraindo dados da OC: {oc_completa}...")
            self.registrar_log(f"Extraindo dados da FSE para a OC: {oc_completa}")
            
            wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@ng-model='vm.search.orderNumber']"))).clear()
            self.driver.find_element(By.XPATH, "//input[@ng-model='vm.search.orderNumber']").send_keys(oc1)
            wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@ng-model='vm.search.orderLine']"))).clear()
            self.driver.find_element(By.XPATH, "//input[@ng-model='vm.search.orderLine']").send_keys(oc2)
            wait.until(EC.element_to_be_clickable((By.ID, "searchBtn"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(@ng-click, 'vm.showFseDetails')]"))).click()
            wait.until(EC.visibility_of_element_located((By.ID, "fseHeader")))
            
            dados = {"OS": os_num}
            
            # Utiliza seletores mais robustos baseados em texto
            oc_item_raw = self.safe_find_text(By.XPATH, "//*[normalize-space()='OC / ITEM']/ancestor::div[1]/following-sibling::div").replace('\n', '/').strip()
            oc_item_split = [x.strip() for x in oc_item_raw.split('/')]
            dados["OC"] = oc_item_split[0] if oc_item_split else ""
            dados["Item"] = oc_item_split[1] if len(oc_item_split) > 1 else ""

            codem_raw = self.safe_find_text(By.XPATH, "//*[normalize-space()='CODEM / DT. REV. ROT.']/ancestor::div[1]/following-sibling::div").replace('\n', '/').strip()
            codem_split = [part.strip() for part in codem_raw.split('/') if part.strip()]
            dados["CODEM"] = codem_split[0] if codem_split else ""
            dados["DT. REV. ROT."] = codem_split[1] if len(codem_split) > 1 else ""

            pn_raw = self.safe_find_text(By.XPATH, "//*[normalize-space()='PN / REV. PN / LID']/ancestor::div[1]/following-sibling::div").replace('\n', ' ').strip()
            pn_parts = [part for part in pn_raw.split() if part]
            dados["PN"] = pn_parts[0] if pn_parts else ""
            dados["REV. PN"] = pn_parts[1] if len(pn_parts) > 1 else ""
            dados["LID"] = pn_parts[2] if len(pn_parts) > 2 else ""
            
            dados["PLANTA"] = self.safe_find_text(By.XPATH, "//*[normalize-space()='PLANTA']/parent::div/following-sibling::div").strip()
            dados["IND. RASTR."] = self.safe_find_text(By.XPATH, "//*[normalize-space()='IND. RASTR.']/ancestor::div[1]/following-sibling::div").replace('\n', '').strip()
            
            seriacao_elements = self.driver.find_elements(By.XPATH, "//b[normalize-space()='NÚMERO DE SERIAÇÃO']/ancestor::div[contains(@class,'border-fse-form-dyn')]//div[contains(@class, 'ng-binding')]")
            dados["NÚMERO DE SERIAÇÃO"] = ", ".join([el.text.strip() for el in seriacao_elements if el.text.strip()])
            
            pn_extraido_match = re.search(r'(\d+-\d+-\d+)', dados.get("PN", ""))
            dados["PN extraído"] = pn_extraido_match.group(1) if pn_extraido_match else "Não encontrado"
            dados["REV. FSE"] = dados.get("REV. PN", "Não encontrada")

            # Otimização: Clica em Voltar ao invés de recarregar a página
            self.find_and_click(wait, ["//button[contains(text(), 'Voltar')]", "//button[contains(text(), 'Back')]"], "Botão Voltar (FSE)")
            wait.until(EC.visibility_of_element_located((By.ID, "searchBtn"))) # Espera a tela de busca recarregar

            return dados
        except Exception:
            self.registrar_log(f"ERRO: Falha ao extrair dados da FSE para a OC {oc_completa}.")
            self.tirar_print_de_erro(oc_completa, "extracao_FSE")
            # Tenta se recuperar voltando para a página de busca
            self.driver.get("https://appscorp2.embraer.com.br/gfs/#/fse/search/1")
            wait.until(EC.visibility_of_element_located((By.ID, "searchBtn")))
            return None
    
    def navegar_para_fse_busca(self, wait):
        self.driver.switch_to.window(self.driver.window_handles[0]) # Garante que está na janela principal
        self.driver.get("https://web.embraer.com.br/irj/portal")
        wait.until(EC.element_to_be_clickable((By.ID, "L2N10"))).click()
        wait.until(EC.number_of_windows_to_be(2))
        for handle in self.driver.window_handles:
            if handle != self.driver.window_handles[0]:
                self.driver.switch_to.window(handle)
                break
        self.prompt_user_action("No navegador, navegue para 'FSE' > 'Busca FSe' e clique em 'Continuar'.")

    def navegar_para_desenhos_engenharia(self, wait):
        self.driver.switch_to.window(self.driver.window_handles[0])
        self.driver.get("https://web.embraer.com.br/irj/portal")
        wait.until(EC.element_to_be_clickable((By.ID, "L2N1"))).click()
        self.prompt_user_action("Valide se a tela 'Desenhos Engenharia' está aberta e clique em 'Continuar'.")
    
    def find_and_click(self, wait, selectors, description):
        for selector in selectors:
            try:
                element = wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                ActionChains(self.driver).move_to_element(element).click().perform()
                return True
            except TimeoutException:
                continue
        self.registrar_log(f"AVISO: Não foi possível clicar no elemento '{description}'.")
        return False

    def buscar_revisao_engenharia(self, wait, part_number):
        if not part_number or part_number == "Não encontrado":
            return "PN não fornecido"
        try:
            self.driver.switch_to.default_content()
            wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "contentAreaFrame")))
            wait.until(EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@id, 'ivuFrm_')]")))

            campo_pn = wait.until(EC.visibility_of_element_located((By.XPATH, "//input[contains(@id, 'PartNumber')]")))
            campo_pn.clear()
            campo_pn.send_keys(part_number)
            self.registrar_log(f"Buscando revisão de engenharia para o PN: {part_number}")
            
            self.find_and_click(wait, ['//*[@id="FOAH.Dplpl049View.cmdGBI"]'], "Botão Desenho")

            # Tenta encontrar a revisão. Se não achar em 5s, parte para o plano B.
            try:
                seletor_rev = '//*[@id="FOAHJJEL.GbiMenu.TreeNodeType1.0.childNode.0.childNode.0.childNode.0.childNode.0-cnt-start"]'
                rev_element = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located((By.XPATH, seletor_rev)))
                revisao = rev_element.text.strip()
                self.registrar_log(f"Revisão de Engenharia encontrada: {revisao}")
                self.find_and_click(wait, ['//*[@id="FOAHJJEL.GbiMenu.cmdRetornarNaveg"]'], "Botão Voltar (Sucesso)")
                return revisao
            except TimeoutException:
                # Se a revisão não apareceu, pode ser um erro de PN não encontrado.
                self.registrar_log(f"AVISO: PN {part_number} não foi encontrado no sistema de Engenharia.")
                self.tirar_print_de_erro(part_number, "busca_revisao_nao_encontrado")
                self.find_and_click(wait, ['//*[@id="FOAH.Dplpl049View.cmdVoltar"]', "//*[contains(@title, 'Voltar')]"], "Botão Voltar (Tela de Erro)")
                return "Não encontrado em ENG"

        except Exception:
            self.registrar_log(f"ERRO CRÍTICO ao buscar revisão para o PN {part_number}.")
            self.tirar_print_de_erro(part_number, "busca_revisao_erro_inesperado")
            # Tenta voltar para a tela principal de qualquer maneira
            try:
                self.find_and_click(WebDriverWait(self.driver, 5), ["//*[contains(@title, 'Voltar')]", '//*[@id="FOAHJJEL.GbiMenu.cmdRetornarNaveg"]'], "Botão Voltar (Genérico)")
            except:
                self.registrar_log("Não foi possível retornar à tela de busca após erro crítico.")
            return "Falha na busca"
        finally:
            # Essencial para sair do iframe e poder interagir com a página principal de novo
            self.driver.switch_to.default_content()

    def safe_find_text(self, by, value):
        try:
            return self.driver.find_element(by, value).text
        except NoSuchElementException:
            return ""

    def tirar_print_de_erro(self, identificador, etapa):
        erros_path = os.path.join(os.getcwd(), ERROS_DIR)
        os.makedirs(erros_path, exist_ok=True)

        identificador_limpo = re.sub(r'[\\/*?:"<>|]', "", str(identificador))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_screenshot = f"erro_{etapa}_{identificador_limpo}_{timestamp}.png"
        screenshot_path = os.path.join(erros_path, nome_screenshot)
        try:
            if self.driver:
                self.driver.save_screenshot(screenshot_path)
                self.registrar_log(f"Screenshot do erro salvo em: '{screenshot_path}'")
        except Exception as e:
            self.registrar_log(f"FALHA AO SALVAR SCREENSHOT: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ValidadorGUI(root)
    root.mainloop()